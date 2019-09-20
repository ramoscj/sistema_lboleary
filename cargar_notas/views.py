# -*- coding: utf-8 -*-
from django.http import HttpResponse, StreamingHttpResponse, HttpResponseRedirect, Http404
from django.contrib.auth.decorators import login_required, permission_required
from django.shortcuts import render
from django.core.exceptions import ObjectDoesNotExist
from io import BytesIO
import xlsxwriter
from .forms import *
from openpyxl import load_workbook
from django.contrib.auth import get_user_model
from xlsxwriter.utility import xl_range_abs, xl_rowcol_to_cell
from django.conf import settings
from notas.models import PdoEscAlum
from docentes.models import MateriaMaestro

import os.path

@login_required
def CargarXlS(request):

	if request.method == 'POST':
		form = CargarXLSForm(request.POST, request.FILES)
		if form.is_valid():
		    if 'archivo' in request.FILES:
		        xls = request.FILES['archivo']    
		        form.handle_uploaded_file(xls)
		        GuardarNotasLapso(xls)
		        return render(request, 'cargar_xls.html',{'form': CargarXLSForm()},)
	else:
		form = CargarXLSForm()
		return render(request, 'cargar_xls.html',{'form': form})

def GuardarNotasLapso(archivo):

	fname = 'C:\Users\Biblioteca\Documents\Notas\lboleary\media\%s' % archivo
	f = os.path.isfile(fname) 
	if f:
		wb = load_workbook(fname)
		celdas = ['F', 'H', 'J', 'L', 'N', 'P', 'R', 'T', 'V', 'X', 'Z', 'AB', 'AD', 'AF', 'AH']
		celdas2 = ['G', 'I', 'K', 'M', 'O', 'Q', 'S', 'U', 'W', 'Y', 'AA', 'AC', 'AE', 'AG', 'AI']
		
		for sheetname in wb.get_sheet_names():
			hoja = wb[sheetname]
			j = 10
			for col in hoja['B10:B59']:
				for cell in col:
					if cell.value is not None:
						try:
							alumno = PdoEscAlum.objects.get(alumno__cedula=cell.value)
							for i in range(len(celdas)):
								nota = hoja[('%s%s') % (celdas[i], j)].value
								if nota is not None:
									valor = hoja[('%s5') % celdas2[i]].value
									descripcion = hoja[('%s5') % celdas[i]].value
									lapso = hoja['D6'].value
									curso = hoja['D7'].value
									docente = hoja['D8'].value
									extra = hoja['D4'].value
									print GuardarNotas(alumno.id, curso, docente, nota, lapso, descripcion, valor, extra)
								i += 1
						except ObjectDoesNotExist:
							pass
					j += 1
		return True
	else:
		return False

def GuardarNotas(alumno, curso, docente, nota, lapso, descripcion, valor, extra):

	try:
		if extra == 1:
			if lapso == 1:
				guardar = PrimeroNTPL(alumno_curso_id=alumno, curso_id=curso, docente_id=docente, nota=nota, descripcion=descripcion, valor=valor)
				guardar.save()
				return True
			if lapso == 2:
				guardar = PrimeroNTSL(alumno_curso_id=alumno, curso_id=curso, docente_id=docente, nota=nota, descripcion=descripcion, valor=valor)
				guardar.save()
				return True
			if lapso == 3:
				guardar = PrimeroNTTL(alumno_curso_id=alumno, curso_id=curso, docente_id=docente, nota=nota, descripcion=descripcion, valor=valor)
				guardar.save()
				return True
		if extra == 2:
			if lapso == 1:
				guardar = SegundoNTPL(alumno_curso_id=alumno, curso_id=curso, docente_id=docente, nota=nota, descripcion=descripcion, valor=valor)
				guardar.save()
				return True
			if lapso == 2:
				guardar = SegundoNTSL(alumno_curso_id=alumno, curso_id=curso, docente_id=docente, nota=nota, descripcion=descripcion, valor=valor)
				guardar.save()
				return True
			if lapso == 3:
				guardar = SegundoNTTL(alumno_curso_id=alumno, curso_id=curso, docente_id=docente, nota=nota, descripcion=descripcion, valor=valor)
				guardar.save()
				return True
		if extra == 3:
			if lapso == 1:
				guardar = TerceroNTPL(alumno_curso_id=alumno, curso_id=curso, docente_id=docente, nota=nota, descripcion=descripcion, valor=valor)
				guardar.save()
				return True
			if lapso == 2:
				guardar = TerceroNTSL(alumno_curso_id=alumno, curso_id=curso, docente_id=docente, nota=nota, descripcion=descripcion, valor=valor)
				guardar.save()
				return True
			if lapso == 3:
				guardar = TerceroNTTL(alumno_curso_id=alumno, curso_id=curso, docente_id=docente, nota=nota, descripcion=descripcion, valor=valor)
				guardar.save()
				return True
		if extra == 4:
			if lapso == 1:
				guardar = CuartoNTPL(alumno_curso_id=alumno, curso_id=curso, docente_id=docente, nota=nota, descripcion=descripcion, valor=valor)
				guardar.save()
				return True
			if lapso == 2:
				guardar = CuartoNTSL(alumno_curso_id=alumno, curso_id=curso, docente_id=docente, nota=nota, descripcion=descripcion, valor=valor)
				guardar.save()
				return True
			if lapso == 3:
				guardar = CuartoNTTL(alumno_curso_id=alumno, curso_id=curso, docente_id=docente, nota=nota, descripcion=descripcion, valor=valor)
				guardar.save()
				return True
		if extra == 5:
			if lapso == 1:
				guardar = QuintoNTPL(alumno_curso_id=alumno, curso_id=curso, docente_id=docente, nota=nota, descripcion=descripcion, valor=valor)
				guardar.save()
				return True
			if lapso == 2:
				guardar = QuintoNTSL(alumno_curso_id=alumno, curso_id=curso, docente_id=docente, nota=nota, descripcion=descripcion, valor=valor)
				guardar.save()
				return True
			if lapso == 3:
				guardar = QuintoNTTL(alumno_curso_id=alumno, curso_id=curso, docente_id=docente, nota=nota, descripcion=descripcion, valor=valor)
				guardar.save()
				return True
	except Exception, e:
		return e

@login_required
def DescargarXls(request, id_docente, id_periodo):

	# create workbook with worksheet
	maestro = MateriaMaestro.objects.filter(maestro=id_docente).filter(periodo=id_periodo)
	
	if maestro.exists():
		output = BytesIO()
		book = xlsxwriter.Workbook(output)
		
		# Formato para titulos - bloqueada
		bold = book.add_format({'bold': True, 'font_name': 'Arial', 'font_size': 11, 'align': 'center', 'valign': 'vcenter', 'locked': True, 'border': 2})
		bold2 = book.add_format({'bold': True, 'font_name': 'Arial', 'font_size': 11, 'align': 'left', 'valign': 'vcenter', 'locked': True, 'border': 2})
		bold_color = book.add_format({'bold': True, 'font_name': 'Arial', 'font_size': 11, 'align': 'center', 'valign': 'vcenter', 'locked': True, 'border': 2, 'bg_color': '#DADADA'})
		# Formato para suma de porcentajes - bloqueada
		format_cort = book.add_format({'bold': True, 'rotation': 90, 'font_name': 'Arial', 'font_size': 9, 'align': 'center', 'valign': 'bottom', 'locked': True , 'border': 2, 'bg_color': '#EEEBEB'})
		format_evl = book.add_format({'bold': True, 'rotation': 90, 'font_name': 'Arial', 'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'locked': False, 'border': 2, 'bg_color': '#EEEBEB'})
		# Formato para titulos de evaluaciones - libre
		format_evl2 = book.add_format({'bold': True, 'rotation': 90, 'font_name': 'Arial', 'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'locked': False, 'border': 2})
		# Formato para lista de alumnos - bloqueada
		format_data = book.add_format({'font_name': 'Arial', 'font_size': 9, 'align': 'left', 'locked': True, 'border': 1})
		format_data2 = book.add_format({'font_name': 'Arial', 'font_size': 9, 'align': 'center', 'locked': True, 'border': 1})
		# Formato para lista de alumnos - libre
		format_data3 = book.add_format({'font_name': 'Arial', 'font_size': 9, 'align': 'left', 'locked': False, 'border': 1})
		# Formato para numeros - bloqeada
		format_num = book.add_format({'num_format': '0.00', 'font_name': 'Arial', 'font_size': 9, 'align': 'center', 'valign': 'bottom', 'locked': True, 'border': 1, 'bg_color': '#EEEBEB'})
		format_des = book.add_format({'num_format': '0', 'font_name': 'Arial', 'font_size': 9, 'align': 'center', 'valign': 'bottom', 'locked': False, 'border': 1})
		format_ent = book.add_format({'num_format': '0', 'font_name': 'Arial', 'font_size': 9, 'align': 'center', 'valign': 'bottom', 'locked': True, 'border': 1, 'bg_color': '#DADADA'})
		format_sum = book.add_format({'num_format': '0.00', 'font_name': 'Arial', 'font_size': 9, 'align': 'center', 'valign': 'bottom', 'locked': True, 'border': 1, 'bg_color': '#EEEBEB'})
		# Para condicion de colores
		format_color1 = book.add_format({'num_format': '0%', 'font_name': 'Arial', 'font_size': 28, 'font_color': 'white', 'align': 'center', 'valign': 'bottom', 'locked': True, 'border': 1, 'bg_color': 'green'})
		format_color2 = book.add_format({'num_format': '0%', 'font_name': 'Arial', 'font_size': 28, 'align': 'center', 'valign': 'bottom', 'locked': True, 'border': 1, 'bg_color': 'yellow'})
		format_color3 = book.add_format({'num_format': '0%', 'font_name': 'Arial', 'font_size': 28, 'font_color': 'white', 'align': 'center', 'valign': 'bottom', 'locked': True, 'border': 1, 'bg_color': 'red'})
		bold3 = book.add_format({'bold': True, 'font_name': 'Arial', 'font_size': 28, 'align': 'center', 'valign': 'vcenter', 'locked': True, 'border': 2})
		# Titulos para las Columnas
		columns = ["Nro", u"Cédula", "Apellidos y Nombres", "Nota"]
		adicional = [ u"Evaluación", 0]

		for carga in maestro:
			sheet = ""
			# Darle nombre a la hoja segun materia, año, seccion y periodo escolar
			sheet = book.add_worksheet(("%s_%s%s")% (carga.materia.nombcorto, carga.seccion.grado, carga.seccion.seccion))
			# Para bloquear hoja y la clave
			sheet.protect('controloleary')
			# Encabezado
			celdas = ['C']
			textos = [u'Código: S0174D0604', u'Año: %s' % carga.seccion.grado, u'Sección: %s' %carga.seccion.seccion, 'Periodo Escolar: %s' % carga.periodo, 'Lapso: I']
			sheet.merge_range('C1:I1', u'LICEO NACIONAL BOLIVARIANO DANIEL FLORENCIO O´LEARY', bold)
			# IDS de los registros
			ids = [carga.seccion.id, carga.seccion.grado, carga.periodo.id, 1]
			# Ingresando los encabezados y los ids modo manual por tener celdas combinadas
			sheet.merge_range('B1:B6', None, bold2)
			sheet.insert_image('B1', 'C:\Users\Biblioteca\Documents\Notas\lboleary/notas\media\imagenes\logo_oleary.png', {'x_scale': 0.25, 'y_scale': 0.25, 'x_offset': 8})
			sheet.write('D2:D2', None, bold2)
			sheet.merge_range('A7:C7', 'Materia: %s' % carga.materia, bold2)
			sheet.merge_range('A8:C8', 'Docente: %s %s' % (carga.maestro.nombres.upper(), carga.maestro.apellidos.upper()) , bold2)
			sheet.write('D3', ids[0] , bold)
			sheet.write('D4', ids[1] , bold)
			sheet.write('D5', ids[2] , bold)
			sheet.write('D6', ids[3] , bold)
			sheet.write('D7', carga.id , bold)
			sheet.write('D8', carga.maestro.id , bold)
			# Ayuda para validar que no se pasen la suma de los % de las evaluaciones
			sheet.merge_range('E2:I4', '=SUM(G5,I5,K5,M5,O5,Q5,S5,U5,W5,Y5,AA5,AC5,AE5,AG5,AI5)/100', bold3)
			sheet.conditional_format('E2:I4', {'type':     'cell',
                                        'criteria': '=',
                                        'value':    '100%',
                                        'format':   format_color1})
			sheet.conditional_format('E2:I4', {'type':     'cell',
                                        'criteria': '<',
                                        'value':    '100%',
                                        'format':   format_color2})
			sheet.conditional_format('E2:I4', {'type':     'cell',
                                        'criteria': '>',
                                        'value':    '100%',
                                        'format':   format_color3})
			sheet.data_validation('E2', {
					'validate': 'any',
					'input_title': u'Suma en "%" de las evaluaciones:',
					'input_message': 'Entre "0%" y "100%"',})

			#Ingresando los encabezados y los ids en bucle
			o = 2
			for texto in textos:
				sheet.write(("%s%s") % (celdas[0], o), texto, bold2)
				o += 1
			
			# Imnnovilizar los paneles
			sheet.freeze_panes('D10')
			# Controlar todas las filas
			row = 8
			# 
			# Titulos de la las filas
			for i,elem in enumerate(columns):
			    sheet.write(row, i, elem, bold)
			    if i == 3:
			    	sheet.write(row, i, elem, bold_color)
			sheet.merge_range('E5:E9', "Corte sin redondeo", format_cort)
			# Titulos de las evaluaciones
			y = 5
			for z in range(15):
				sheet.merge_range(4, y, 8, y, "%s %s" % (adicional[0], z+1), format_evl2)
				# Mensaje de ayuda
				sheet.data_validation(4, y, 8, y, {
					'validate': 'any',
					'input_title': 'IMPORTANTE:     ',
					'input_message': u'Ingrese una breve descripción de la evaluación'})
				y += 1
				sheet.merge_range(4, y, 8, y, adicional[1], format_evl)
				# Mensaje de ayuda
				sheet.data_validation(4, y, 8, y, {
					'validate': 'integer',
					'criteria': 'between',
					'minimum': 0,
					'maximum': 20,
					'input_title': u'Valor en "%" de la evaluación:',
					'input_message': 'Entre 1 y 20',
					'error_title': 'Valor invalido',
					'error_message': 'Intente nuevamente, rango del 1 al 20.',
					'ignore_blank': False})
				y += 1
			row += 1
			# Ajustando ancho de las columnas
			sheet.set_column('A8:A8', 4)
			sheet.set_column('B8:B8', 15)
			sheet.set_column('C8:C8', 35)
			sheet.set_column('D8:D8', 5)
			sheet.set_column('E:AI', 5)
			sheet.set_row(8, 26)
			# Ingresando los datos de los alumnos
			x = 1
			alumnos = PdoEscAlum.objects.filter(periodo=carga.periodo.id).filter(asignado=carga.seccion.id)
			for consulta in alumnos:
				sheet.write(row, 0, x, format_data2)
				sheet.write(row, 1, consulta.alumno.cedula, format_data)
				sheet.write(row, 2, ("%s %s") % (consulta.alumno.apellidos.upper(), consulta.alumno.nombres.upper()), format_data)
				# Ingresando las formulas para calculos de las notas
				j = 6
				data = ""
				for k in range(15):
					rango1 = xl_rowcol_to_cell(4, j, row_abs=True, col_abs=True)
					rango2 = xl_rowcol_to_cell(row, j-1, row_abs=True, col_abs=True)
					# Agregando bordes a las celdas de las evaluaciones
					sheet.write(row, j-1, None, format_des)
					sheet.data_validation(row, j-1, row, j-1, {
						'validate': 'integer',
						'criteria': 'between',
						'minimum': 1,
						'maximum': 20,
						'error_title': 'Valor invalido',
						'error_message': 'Intente nuevamente, rango del 1 al 20.',
						'ignore_blank': False})
					# Formula para calcular el porcentaje de la evaluacion
					sheet.write_formula(row, j, '=(%s*%s)/100' % (rango1,rango2), format_num)
					# Recoger las casillas para la suma de los porcentajes
					foco = xl_rowcol_to_cell(row, j, row_abs=True, col_abs=True)
					data = ("%s,") % foco + data
					
					k += 1
					j += 2
				# Formula para nota final - numero entero
				redondear = xl_rowcol_to_cell(row, 4, row_abs=True, col_abs=True)
				sheet.write(row, 3, '=ROUND(%s,0)' % redondear, format_ent)
				# Formula para sumar los porcentajes
				sheet.write_formula(row, 4, '=SUM(%s)' % data, format_sum)
				x += 1
				row += 1
			# Para rellenear el formato hasta llegar a los 50
			while x <= 50:
				sheet.write(row, 0, x, format_data2)
				sheet.write(row, 1, None, format_data3)
				sheet.write(row, 2, None, format_data3)
				# Ingresando las formulas para calculos de las notas
				j = 6
				data = ""
				for k in range(15):
					rango1 = xl_rowcol_to_cell(4, j, row_abs=True, col_abs=True)
					rango2 = xl_rowcol_to_cell(row, j-1, row_abs=True, col_abs=True)
					# Agregando bordes a las celdas de las evaluaciones
					sheet.write(row, j-1, None, format_des)
					sheet.data_validation(row, j-1, row, j-1, {
						'validate': 'integer',
						'criteria': 'between',
						'minimum': 1,
						'maximum': 20,
						'error_title': 'Valor invalido',
						'error_message': 'Intente nuevamente, rango del 1 al 20.',
						'ignore_blank': False})
					# Formula para calcular el porcentaje de la evaluacion
					sheet.write_formula(row, j, '=(%s*%s)/100' % (rango1,rango2), format_num)
					# Recoger las casillas para la suma de los porcentajes
					foco = xl_rowcol_to_cell(row, j, row_abs=True, col_abs=True)
					data = ("%s,") % foco + data
					
					k += 1
					j += 2
				# Formula para nota final - numero entero
				redondear = xl_rowcol_to_cell(row, 4, row_abs=True, col_abs=True)
				sheet.write(row, 3, '=ROUND(%s,0)' % redondear, format_ent)
				# Formula para sumar los porcentajes
				sheet.write_formula(row, 4, '=SUM(%s)' % data, format_sum)
				x += 1
				row += 1
			salidaxls = ("CARGA_ACADEMICA_%s_%s_%s") % ((carga.maestro.nombres[0:3].upper(), carga.maestro.apellidos[0:3].upper(), carga.periodo))

		book.close()  # close book and save it in "output"
		output.seek(0)  # seek stream on begin to retrieve all data from it

		# send "output" object to stream with mimetype and filename
		response = StreamingHttpResponse(
		    output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
		response['Content-Disposition'] = 'attachment; filename=%s.xlsx' % salidaxls
		return response

	else: 
		raise Http404("El Docente No Existe!")
